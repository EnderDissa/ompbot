# -*- coding: utf-8 -*-
import json
import os
from datetime import datetime

import vk_api
from random import randint as rd

from openpyxl.styles import PatternFill, Side, Border
from vk_api import VkApi
from vk_api.bot_longpoll import VkBotLongPoll, VkBotEventType, VkBotMessageEvent
from vk_api.keyboard import VkKeyboard, VkKeyboardColor
import pandas as pd
import requests
import traceback
from datetime import datetime as date
import openpyxl
import re


def lsend(id, text):
    print("sended to " + str(id))
    vk_session.method('messages.send', {'user_id': id, 'message': text, 'random_id': 0})


def lsend_withA(id, text, attachment):
    print("sended to " + str(id))
    vk_session.method('messages.send', {'user_id': id, 'message': text, 'attachment': attachment, 'random_id': 0})


def send(id, text):
    print("sended to " + str(id))
    vk_session.method('messages.send', {'chat_id': id, 'message': text, 'random_id': 0})


def send_withA(id, text, attachment, title, sender, kolgost=0):
    print("sended to " + str(id))
    keyboard = vk_api.keyboard.VkKeyboard(inline=True)
    keyboard.add_callback_button(label="ОТПРАВИТЬ", payload={"type": "send", 'sender': sender, 'title': title},
                                 color=VkKeyboardColor.SECONDARY)
    keyboard.add_callback_button(label="СОГЛАСОВАТЬ", payload={"type": "approve", 'sender': sender, 'title': title, 'isSended': False},
                                 color=VkKeyboardColor.POSITIVE)
    keyboard = keyboard.get_keyboard()
    vk_session.method('messages.send',
                      {'chat_id': id, 'message': text, 'attachment': attachment, 'keyboard': keyboard, 'random_id': 0})


def editkb(peer_id, cmid, type):
    keyboard = vk_api.keyboard.VkKeyboard(inline=True)
    keyboard.add_callback_button(label="ОТПРАВЛЕНО", payload={"type": "sended", 'sender': sender, 'title': title},
                                 color=VkKeyboardColor.NEGATIVE)
    keyboard.add_callback_button(label=("СОГЛАСОВАНО" if type == "approve" else "СОГЛАСОВАТЬ"),
                                 payload={"type": ("approved" if type == "approve" else "approve"), 'sender': sender,
                                          'title': title, 'isSended':(False if type=="send" else True)},
                                 color=(VkKeyboardColor.NEGATIVE if type == "approved" else VkKeyboardColor.POSITIVE))
    keyboard = keyboard.get_keyboard()

    original_message = vk.messages.getById(
        peer_id=peer_id,
        cmids=cmid)
    original_text = original_message['items'][0]['text']
    original_attachment = original_message['items'][0]['attachments'][0]['doc']
    original_attachment = "doc" + str(original_attachment['owner_id']) + '_' + str(original_attachment['id'])

    vk.messages.edit(peer_id=peer_id, conversation_message_id=cmid, keyboard=keyboard, message=original_text,
                     attachment=original_attachment)


def sender(sender_type):
    pass


def attachment_extract(url, name):
    response = requests.get(url)
    if not os.path.exists('xlsx/' + name):
        dir = 'xlsx/' + name
        os.mkdir(dir)
        print("новый клуб: " + name)
    path = "xlsx/" + name + "/" + ("_".join(str(date.now())[:-7].replace(":", "-").split())) + ".xlsx"
    with open(path, "wb") as f:
        f.write(response.content)
        return path


def check_excel(path):
    row = []
    rows = []
    data = openpyxl.load_workbook(path)
    sheet = data.active
    correct_meta = ['Корпус:', '№', 'Фамилия', 'Дата, время:', 'Имя', 'Отчество', 'Название мероприятия:',
                    'Серия и номер паспорта', 'Номер телефона', 'Ответственный от подразделения:',
                    'Ежков Павел Игроевич, заместитель директора ОМП', 79213422059, 'Контактное лицо:']
    meta = [sheet['A1'].value, sheet['A2'].value, sheet['B2'].value, sheet['C1'].value, sheet['C2'].value,
            sheet['D2'].value, sheet['E1'].value, sheet['E2'].value, sheet['F2'].value, sheet['G1'].value,
            sheet['G2'].value, sheet['G3'].value, sheet['H1'].value]
    korpus = sheet['B1'].value
    date_time = str(sheet['D1'].value)
    name = sheet['F1'].value
    rukovod = sheet['H2'].value
    rukovod_phone = sheet['H3'].value

    date = date_time.split()[0]

    if correct_meta == meta:
        if date_time=="01.01.2025  09:00-23:00" or "Шаблон" in name or "Шаблон" in rukovod or rukovod_phone==79633336075 or rukovod_phone=="79633336075":
            return "01", rows
        i = 0;j=0
        cyrillic_lower_letters = 'абвгдеёжзийклмнопрстуфхцчшщъыьэюя'
        while True:
            j += 1
            col = str(j)
            if sheet['A' + col].value is None: break
        lenrow=j-3
        while True:
            i += 1
            col = str(i)
            if sheet['A' + col].value is None: break
            row = [sheet['A' + col].value, sheet['B' + col].value.strip(), sheet['C' + col].value.strip(),
                   str(sheet['D' + col].value).strip(), str(sheet['E' + col].value).strip().zfill(10),
                   str(sheet['F' + col].value).strip(), sheet['G' + col].value, sheet['H' + col].value]

            if i < 3:
                rows.append(row)
                continue
            if row[0] != i - 2: return "A" + col
            for _ in row[1].lower():
                if _ not in cyrillic_lower_letters: return "B" + col
            row[1]=row[1][0].upper()+row[1][1:].lower()
            for _ in row[2].lower():
                if _ not in cyrillic_lower_letters: return "C" + col + _
            row[2] = row[2][0].upper() + row[2][1:].lower()
            for _ in str(row[3]).lower():
                if _ not in cyrillic_lower_letters: return "D" + col
            row[3] = row[3][0].upper() + row[3][1:].lower()
            if not (row[4].isdigit() or not(re.findall(r"\d{10}",row[4]))) or row[4][:2]=='00': return "E" + col
            if not (row[5].isdigit()): return "F" + col
            digits = re.findall(r"7\d{10}", row[5])[0]
            if not digits: return "F" + col
            if lenrow>2: nomer="8-"+digits[1:4]+"-"+digits[4:7]+"-"+digits[7:9]+"-"+digits[9:] #8-xxx-xxx-xx-xx
            else: nomer=digits
            row[5]=nomer
            rows.append(row)
    else:
        return "00", rows
    return "success", rows


def create_excel(path, rows):
    data = openpyxl.Workbook()
    sheet = data.active

    sheet.title = "Согласование СЗ"
    fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    border = Border(
        left=Side(border_style="medium", color='FF000000'),
        right=Side(border_style="medium", color='FF000000'),
        top=Side(border_style="medium", color='FF000000'),
        bottom=Side(border_style="medium", color='FF000000'),
        diagonal=Side(border_style="medium", color='FF000000'),
        diagonal_direction=0,
        outline=Side(border_style="medium", color='FF000000'),
        vertical=Side(border_style="medium", color='FF000000'),
        horizontal=Side(border_style="medium", color='FF000000')
    )

    for i in range(len(rows)):
        for j in range(len(rows[i])):
            cell = sheet.cell(row=i + 1, column=j + 1)
            cell.value = rows[i][j]
            if cell.value:
                cell.border = border

    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column_letter].width = adjusted_width

    sheet["A1"].fill = fill
    sheet["C1"].fill = fill
    sheet["E1"].fill = fill
    sheet["G1"].fill = fill
    sheet["H1"].fill = fill
    sheet["A2"].fill = fill
    sheet["B2"].fill = fill
    sheet["C2"].fill = fill
    sheet["D2"].fill = fill
    sheet["E2"].fill = fill
    sheet["F2"].fill = fill
    sheet["G2"].fill = fill
    sheet["G3"].fill = fill
    try:
        data.save(path)
        return True
    except:
        print("ERROR")
        return False


with open("token.txt", 'r') as f:
    token = f.readline()

vk_session = vk_api.VkApi(token=token)
vk = vk_session.get_api()

enote = 'https://ursi.yonote.ru/share/clubs/doc/sluzhebnye-zapiski-i-prohod-gostej-bihQHvmk8w'
groupid = 228288169
admin = [297002785]
ignore = []
longpoll = VkBotLongPoll(vk_session, groupid)

print("работай")

while True:
    try:
        for event in longpoll.listen():
            if event.type == VkBotEventType.MESSAGE_EVENT:
                pl = event.object.get('payload')
                if pl:
                    conversation_message_id = event.object['conversation_message_id']
                    peer_id = event.object['peer_id']

                    type = pl['type']
                    sender = int(pl['sender'])
                    title = pl['title']

                    tts = "Ваша служебная записка " + title

                    if type == "send":
                        tts += "\n принята и отправлена на согласование!"
                        editkb(peer_id=peer_id, cmid=conversation_message_id, type="sended")
                    elif type == "approve":
                        is_sended = pl['isSended']
                        if is_sended:
                            tts += "\nсогласована и внесена в систему для отображения на мониторе охраны!"
                        else:
                            tts += "\nсогласована и внесена в систему для получения QR на терминале!"
                        editkb(peer_id=peer_id, cmid=conversation_message_id, type="approved")
                    else:
                        continue

                    lsend(sender, tts)
            if event.type == VkBotEventType.MESSAGE_NEW:
                tts = ''

                time = int(str(date.now().time())[:2])
                weekday = date.today().weekday()
                month = int(str(date.now().date())[-5:-3])
                day = int(str(date.now().date())[-2:])
                if (month==12 and day>=28) or (month==1 and day<=8):
                    tts += "С новым годом! Служебные записки не согласуются на каникулах. Вы можете отправить документ, " \
                           "бот его обработает, но согласование получите только после 9 января. Если " \
                           "ситуация срочная, пишите \"МЕНЕДЖЕР\"\n\n"
                elif weekday > 4:
                    tts += "Внимание! Служебные записки не согласуются по выходным. Вы можете отправить документ, " \
                           "бот его обработает, но согласование получите только в понедельник. Если " \
                           "ситуация срочная, пишите \"МЕНЕДЖЕР\"\n\n"
                elif weekday == 4 and time >= 16:
                    tts += "Внимание! По пятницам служебные записки не согласуются после 16:00. Вы можете отправить " \
                           "документ, бот его обработает, но согласование получите только в понедельник. Если " \
                           "ситуация срочная, пишите \"МЕНЕДЖЕР\"\n\n"
                elif time >= 17:
                    tts += "Внимание! Служебные записки не согласуются после 17:00. Вы можете отправить документ, " \
                           "бот его обработает, но согласование получите только завтра. Если " \
                           "ситуация срочная, пишите \"МЕНЕДЖЕР\"\n\n"
                elif time <10:
                    tts += "Внимание! Служебные записки не обрабатываются до 10:00. Вы можете отправить документ, " \
                           "бот его обработает, но согласование получите только в рабочее время. Если " \
                           "ситуация срочная, пишите \"МЕНЕДЖЕР\"\n\n"

                if event.from_chat:
                    id = event.chat_id
                    uid = event.obj['message']['from_id']
                    peer_id = 2000000000 + uid


                else:
                    uid = event.message.from_id

                    user_get = vk.users.get(user_ids=(uid))
                    user_get = user_get[0]
                    uname = user_get['first_name']
                    usurname = user_get['last_name']

                    peer_id = 2000000000 + uid
                    msg = event.object.message['text'].lower()
                    msgs = msg.split()
                    attachment = event.object.message['attachments']
                    if not attachment:
                        if vk_session.method('groups.isMember', {'group_id': groupid, 'user_id': uid}) == 0:
                            tts += "Бот создан для предобработки служебных записок в университете ИТМО и доступен только клубам. Поэтому чтобы иметь доступ к обработке служебных записок необходимо подписаться на это сообщество, ссылку ты можешь найти в еноте или спросить в группе тг!\n\nПосле подписки отправь ещё одно сообщение. Только в случае возникновения проблем пиши \"МЕНЕДЖЕР\""
                        else:
                            tts += "Отправь мне служебную записку, я проведу предпроверку. Если всё хорошо, я отправлю её на обработку, после чего жди сообщения от менеджера. Если возникла проблема, пиши \"МЕНЕДЖЕР\"\nP.S. обязательно отправляй служебные записки в формате, указанном в yonote: " + enote

                    if msgs:
                        if uid in admin:
                            if msgs[0] == "stop":
                                exit()
                            elif msgs[0] == "sender":
                                sender(msgs[1])
                                tts = "готово"

                    if "менеджер" in msg or "админ" in msg:
                        if uid in ignore:
                            ignore.remove(uid)
                            tts = "Надеюсь, вопрос снят!"
                            send(1, uname + " " + usurname + " больше не вызывает. прямая ссылка:\nvk.com/gim" + str(
                                groupid) + "?sel=" + str(uid))
                            continue
                        else:
                            ignore.append(uid)
                            tts = "Принято, сейчас позову! Напиши свою проблему следующим сообщением. Когда вопрос будет решён, напиши команду ещё раз."
                            send(1, uname + " " + usurname + " вызывает! прямая ссылка:\nvk.com/gim" + str(
                                groupid) + "?sel=" + str(uid))
                        lsend(uid, tts)

                    if uid in ignore:
                        continue
                    else:
                        attachment = event.object.message['attachments']

                        if attachment:
                            attachment = attachment[0]
                            attachment = attachment['doc'] if attachment['type'] == 'doc' else None
                        else:
                            None

                        if attachment:
                            attachment_title = attachment['title']
                            attachment_ext = attachment['ext']
                            attachment_url = attachment['url']
                            if (not (re.match(r'СЗ_[а-яёА-ЯЁa-zA-Z]+\.', attachment_title))) or ("шаблон" in attachment_title and uid not in admin):
                                tts += "ошибка в названии файла. пример:\nСЗ_шаблон.xlsx\nдопускается:\nСЗ_шаблон.метаинф.xlsx\nВместо \"шаблон\" везде название клуба (без пробелов)."
                                lsend(uid, tts)
                                continue
                            attachment_title = re.search(r'СЗ_[а-яёА-ЯЁa-zA-Z]+\.', attachment_title).group()[3:]

                            path = attachment_extract(attachment_url, attachment_title)

                            check = check_excel(path)
                            if check[0] == "success":
                                rows = check[1]
                                tts += "Принято! Отправил на проверку, ожидайте ответа."
                                newname = "СЗ_"+attachment_title[:attachment_title.find(".")] + "_" + "_".join(
                                    rows[0][3].replace(":", "-").replace(".", "-").split())
                                newpath = newname + ".xlsx"
                                create_excel(newpath, rows)

                                result = json.loads(requests.post(
                                    vk.docs.getMessagesUploadServer(type='doc',
                                                                    peer_id=event.object.message['peer_id'])[
                                        'upload_url'],
                                    files={'file': open(newpath, 'rb')}).text)
                                jsonAnswer = vk.docs.save(file=result['file'], title=newname, tags=[])
                                attachment = f"doc{jsonAnswer['doc']['owner_id']}_{jsonAnswer['doc']['id']}"

                                lsend_withA(uid, tts, attachment)
                                kolgost=check[1][-1][0]
                                korpus=check[1][0][1]
                                data=check[1][0][3]
                                merotitle=check[1][0][5]
                                org=check[1][1][7]
                                orgnomer=str(check[1][2][7])
                                send_withA(1, "новая проходка: vk.com/gim" + str(groupid) + "?sel=" + str(uid)+ "\nотправитель: " +uname+" "+usurname+"\nорганизатор: "+org+"("+orgnomer+")"+"\nназвание мероприятия: "+merotitle+"\nкорпус: "+korpus+"\nдата: "+data+"\nколичество гостей: "+str(kolgost),
                                           attachment, newpath, uid, int(kolgost))

                                continue
                            elif check[0] == "00":
                                tts += "ошибка в одной из ячеек, которые нельзя менять. перепроверьте A1, A2, B2, C1, C2, D2, E1, E2, F2, G1, G2, G3, H1 по шаблону"
                            elif check[0]=="01":
                                tts+="ошибка в одной из ячеек, которые необходимо было изменить. поменяйте шаблон!"
                            else:
                                tts += "ошибка в ячейке " + check

                    lsend(uid, tts)
    except Exception:
        traceback.print_exc()
