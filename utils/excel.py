import re
import openpyxl
from openpyxl.styles import PatternFill, Side, Border
from datetime import datetime, date 


def check_excel(path):
    rows = []
    data = openpyxl.load_workbook(path)
    sheet = data.active
    correct_meta = ['Корпус:', '№', 'Фамилия', 'Дата, время:', 'Имя', 'Отчество', 'Название мероприятия:',
                    'Серия и номер паспорта', 'Номер телефона', 'Ответственный от подразделения:',
                    'Калугина Анна Владимировна, ведущий менеджер ОМП', 79514373833, 'Контактное лицо:']
    meta = [sheet['A1'].value, sheet['A2'].value, sheet['B2'].value, sheet['C1'].value, sheet['C2'].value,
            sheet['D2'].value, sheet['E1'].value, sheet['E2'].value, sheet['F2'].value, sheet['G1'].value,
            sheet['G2'].value, sheet['G3'].value, sheet['H1'].value]
    korpus = sheet['B1'].value
    date_time = str(sheet['D1'].value)
    name = sheet['F1'].value
    rukovod = sheet['H2'].value
    rukovod_phone = sheet['H3'].value

    correct_meta_otv = ['Романова Софья Александровна, директор ОМП', 79650431766]
    #correct_meta_otv = ['Калугина Анна Владимировна, ведущий менеджер ОМП', 79514373833]

    date_str = date_time.split()[0]
    now = date.today()
    if correct_meta == meta:
        if date_time == "01.01.2025  09:00-23:00" or "Шаблон" in name or "Шаблон" in rukovod or rukovod_phone == 79633336075 or rukovod_phone == "79633336075":
            return "01", rows
        date_time = datetime.strptime(str(date_time.split()[0]), "%d.%m.%Y").date()
        if date_time < now:
            return "02", rows
        i = 0
        j = 0
        cyrillic_lower_letters = 'абвгдеёжзийклмнопрстуфхцчшщъыьэюя- '
        while True:
            j += 1
            col = str(j)
            if sheet['A' + col].value is None: break
        lenrow = j - 3
        while True:
            i += 1
            col = str(i)
            if sheet['A' + col].value is None: break
            row = [sheet['A' + col].value, sheet['B' + col].value.strip(), sheet['C' + col].value.strip(),
                   str(sheet['D' + col].value).strip(), (str(sheet['E' + col].value).replace(" ", "").zfill(10) if i>3 else str(sheet['E' + col].value)),
                   str(sheet['F' + col].value).strip(), sheet['G' + col].value, sheet['H' + col].value]

            if i < 3:
                if i == 2: row[6] = correct_meta_otv[0]
                rows.append(row)
                continue

            if i == 3:

                digits = re.findall(r"7\d{10}", str(int(float(row[6]))))[0]
                row[6] = correct_meta_otv[1]
                digits = re.findall(r"7\d{10}", str(int(float(row[7]))))[0]
                row[7] = digits


            if row[0] != i - 2: return "A" + col
            for _ in row[1].lower():
                if _ not in cyrillic_lower_letters: return "B" + col
            row[1] = row[1][0].upper() + row[1][1:].lower()
            for _ in row[2].lower():
                if _ not in cyrillic_lower_letters: return "C" + col + _
            row[2] = row[2][0].upper() + row[2][1:].lower()
            if str(row[3]) != "None":
                for _ in str(row[3]).lower():
                    if _ not in cyrillic_lower_letters:
                        return "D" + col
                row[3] = row[3][0].upper() + row[3][1:].lower()
            else:
                row[3] = ""

            if not ((row[4].isdigit() or row[4].replace(".", "", 1).isdigit()) or not (
            re.findall(r"\d{10}", row[4]))) or row[4][:2] == '00': return "E" + col
            digits = re.findall(r"\d{10}", str(int(float(row[4]))).zfill(10))[0]
            row[4] = digits
            if not (row[5].isdigit() or row[5].replace(".", "", 1).isdigit()) or not (
            re.findall(r"7\d{10}", row[5])): return "F" + col
            digits = re.findall(r"7\d{10}", str(int(float(row[5]))))[0]
            if not digits: return "F" + col
            if lenrow > 2:
                nomer = "8-" + digits[1:4] + "-" + digits[4:7] + "-" + digits[7:9] + "-" + digits[9:]  # 8-xxx-xxx-xx-xx
            else:
                nomer = digits
            row[5] = nomer
            rows.append(row)
    else:
        return "00", rows
    return "success", rows


def create_excel(path, rows):
    data = openpyxl.Workbook()
    sheet = data.active

    sheet.title = "Согласование СЗ"
    fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    border = Border(
        left=Side(border_style="medium", color='FF000000'),
        right=Side(border_style="medium", color='FF000000'),
        top=Side(border_style="medium", color='FF000000'),
        bottom=Side(border_style="medium", color='FF000000'),
        diagonal=Side(border_style="medium", color='FF000000'),
        diagonal_direction=0,
        outline=Side(border_style="medium", color='FF000000'),
        vertical=Side(border_style="medium", color='FF000000'),
        horizontal=Side(border_style="medium", color='FF000000')
    )

    for i in range(len(rows)):
        for j in range(len(rows[i])):
            cell = sheet.cell(row=i + 1, column=j + 1)
            cell.value = rows[i][j]
            if cell.value:
                cell.border = border

    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column_letter].width = adjusted_width

    sheet["A1"].fill = fill
    sheet["C1"].fill = fill
    sheet["E1"].fill = fill
    sheet["G1"].fill = fill
    sheet["H1"].fill = fill
    sheet["A2"].fill = fill
    sheet["B2"].fill = fill
    sheet["C2"].fill = fill
    sheet["D2"].fill = fill
    sheet["E2"].fill = fill
    sheet["F2"].fill = fill
    sheet["G2"].fill = fill
    sheet["G3"].fill = fill
    try:
        data.save(path)
        return True
    except:
        print("ERROR")
        return False
